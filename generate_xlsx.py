"""
AdRide Unit Economics — generator XLSX v2
Poprawki vs v1:
  - Formuły: brak zagnieżdżonych = (root cause błędów Excel)
  - Koszt prawny: jednorazowy koszt startu, nie amortyzowany miesięcznie
  - Scroll bary: dodane przez win32com po wygenerowaniu
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, zipfile, shutil

wb = openpyxl.Workbook()

PLN   = '#,##0 "zł"'
PCT   = '0.0"%"'
NUM   = '#,##0'
NUM2  = '0.00'
REF_CAR = 200000

# ──────────────────────────────────────────────────────────────
# STYLE HELPERS
# ──────────────────────────────────────────────────────────────
def fill(c): return PatternFill('solid', fgColor=c)
def fnt(bold=False, color='000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')
def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def brd(color='CED4DA'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, value=None, bg=None, fg='000000', bold=False,
       fmt=None, h='left', italic=False, size=10, border=True, wrap=False):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value = value
    if bg: c.fill = fill(bg)
    c.font = fnt(bold, fg, size, italic)
    c.alignment = aln(h, 'center', wrap)
    if fmt: c.number_format = fmt
    if border: c.border = brd()
    return c

def hdr(ws, row, cols, bg='0F3460', fg='FFFFFF', size=10):
    for i, v in enumerate(cols, 1):
        sc(ws, row, i, v, bg=bg, fg=fg, bold=True, h='center', size=size)
    ws.row_dimensions[row].height = 18

def sec(ws, row, title, ncols, bg='E94560', fg='FFFFFF'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1)
    c.value = f"  {title}"
    c.fill = fill(bg); c.font = fnt(True, fg, 10)
    c.alignment = aln('left', 'center'); c.border = brd(bg)
    ws.row_dimensions[row].height = 20

# ──────────────────────────────────────────────────────────────
# SHEET 1 — DANE WEJŚCIOWE
# ──────────────────────────────────────────────────────────────
ws_in = wb.active
ws_in.title = "Dane wejściowe"
ws_in.sheet_view.showGridLines = False
for col, w in zip('ABCDEF', [36, 20, 28, 16, 14, 6]): ws_in.column_dimensions[col].width = w

ws_in.merge_cells('A1:E1')
c = ws_in['A1']
c.value = "🚗  AdRide — Unit Economics Calculator"
c.fill = fill('0F3460'); c.font = Font(bold=True, color='FFFFFF', size=14, name='Calibri')
c.alignment = aln('center'); ws_in.row_dimensions[1].height = 32

ws_in.merge_cells('A2:E2')
c = ws_in['A2']
c.value = "Zmień żółte komórki (lub użyj suwaków) — wyniki aktualizują się automatycznie we wszystkich arkuszach"
c.fill = fill('FFF3CD'); c.font = Font(color='7D6608', size=10, italic=True, name='Calibri')
c.alignment = aln('center'); ws_in.row_dimensions[2].height = 18

# Inputs definition: (label, key, default, min_v, max_v, step, scale, comment, section_before)
#   scale: Excel scrollbar is int → actual_value = scrollbar_value * scale
INPUTS = [
    # section
    (None, None, None, None, None, None, None, None, "🚘 FLOTA & EKSPOZYCJA"),
    ("Liczba aut",           "num_cars",   10,  1,   100,  1,    1,     "aut (1–100)"),
    ("Kilometry / auto / mies.", "km_per_month", 1000, 200, 10000, 100, 100, "km — rzeczywista ekspozycja GPS"),
    ("% km pozamiejskich",   "highway_share", 30, 0, 100, 5, 5, "% km drogi/jazda >50 km/h (taniej, mniej ekspozycji)"),
    ("Wartość auta",         "car_value",  200000, 50000, 600000, 10000, 10000, "zł — referencja: 200 000 zł"),
    ("Czas kampanii",        "months",     6,   1,   24,   1,    1,     "miesięcy"),

    (None, None, None, None, None, None, None, None, "👤 KIEROWCA (stawki miejska / pozamiejska)"),
    ("Stawka kierowcy / km miejski", "driver_rate_per_km", 0.60, 0.10, 2.00, 0.05, 0.05, "zł/km — teren zabud. ≤50 km/h"),
    ("Stawka kierowcy / km pozamiejski", "driver_rate_highway", 0.25, 0.05, 1.00, 0.05, 0.05, "zł/km — drogi/jazda >50 km/h"),
    ("Współczynnik oklejenia", "wrap_factor", 1.0, None, None, None, None, "1.0=całe, 0.5=pół, 0.3=naklejki"),

    (None, None, None, None, None, None, None, None, "📦 MODEL — ABONAMENTOWY (AdRide org. oklejenie, min. 12 mc)"),
    ("Koszt oklejenia / auto",  "wrap_cost",   8000, 2000, 15000, 500, 500, "zł netto"),
    ("Koszt demontażu / auto",  "unwrap_cost", 1200, 500,  3000,  100, 100, "zł netto"),
    ("Narzut ABO vs SUB (%)",   "abo_uplift",  85, 0, 250, 5, 5, "% — kryje oklejenie + logistykę (ABO droższy)"),

    (None, None, None, None, None, None, None, None, "💰 PRICING (per-km, miejska / pozamiejska)"),
    ("Stawka reklamodawcy / km miejski", "adv_rate_per_km", 1.20, 0.30, 4.00, 0.10, 0.10, "zł/km — teren zabud. ≤50 km/h"),
    ("Stawka reklamodawcy / km pozamiejski", "adv_rate_highway", 0.50, 0.10, 2.00, 0.10, 0.10, "zł/km — drogi/jazda >50 km/h"),
    ("Współczynnik prestiżu",    "prestige_coeff",  0.30, 0.0, 0.40, 0.05, 0.05, "premium za wartość auta (symetryczny)"),
    ("Marża docelowa (%)",                "target_margin_pct", 30, 5,  70,    1,   1,   "% (próg alertu)"),

    (None, None, None, None, None, None, None, None, "🔒 KOSZTY"),
    ("Jednorazowy audyt prawny",  "legal_cost", 15000, 5000, 50000, 1000, 1000, "zł netto — koszt startu, NIE miesięczny"),
    ("Koszty ops. platformy / mies.", "ops_cost", 2000, 500, 10000, 250, 250, "zł/mies."),
]

input_cells = {}   # key -> (row, col=2)
# scroll_config: key -> (min_sb, max_sb, scale)  — for scrollbars
scroll_config = {}

row = 3
for item in INPUTS:
    if item[0] is None:  # section header
        sec(ws_in, row, item[8], 5)
        row += 1
        hdr(ws_in, row, ["Parametr", "Wartość", "Uwagi", "↕", ""], bg='1A2744', size=9)
        row += 1
        continue

    label, key, default, minv, maxv, step, scale, comment = item[:8]
    ws_in.row_dimensions[row].height = 22

    sc(ws_in, row, 1, label, bg='EEF2FF')

    # Value cell
    c = ws_in.cell(row=row, column=2)
    c.fill = fill('FFF3CD')
    c.font = fnt(True, '000000', 11)
    c.alignment = aln('right', 'center')
    c.border = brd('E94560')
    if isinstance(default, float):  c.number_format = '0.00'
    elif default > 100:             c.number_format = PLN
    else:                           c.number_format = NUM
    input_cells[key] = (row, 2)

    if scale is not None and scale != 1:
        # Helper cell F: integer (value/scale) linked to scroll bar; B reads F*scale
        # scale may be fractional (np. 0.05 dla stawek/km) → round() chroni przed błędem float
        fh = ws_in.cell(row=row, column=6)
        fh.value = round(default / scale)
        fh.number_format = NUM
        fh.border = brd()
        c.value = f"=F{row}*{scale}"
    else:
        c.value = default

    # Comment
    sc(ws_in, row, 3, comment, bg='FFFFFF', italic=True, fg='666666', size=9)

    # Scrollbar placeholder (col D = small label, col E = where COM will place the control)
    if minv is not None and scale is not None:
        min_sb = round(minv / scale)
        max_sb = round(maxv / scale)
        val_sb = round(default / scale)
        scroll_config[key] = (row, min_sb, max_sb, scale, val_sb)
        sc(ws_in, row, 4, "↕", bg='F0F4FF', fg='0969DA', h='center', size=9, bold=True)
    else:
        sc(ws_in, row, 4, "", bg='FFFFFF')

    sc(ws_in, row, 5, "", bg='FFFFFF')
    row += 1

# ──────────────────────────────────────────────────────────────
# FORMULA BUILDING BLOCKS (pure expressions, no leading =)
# Build each expression fresh from inp() references to avoid nested = bug
# ──────────────────────────────────────────────────────────────
def inp(key):
    """Return Excel cell reference for input key."""
    r, c = input_cells[key]
    return f"'Dane wejściowe'!{get_column_letter(c)}{r}"

def xprestige():
    "Prestige multiplier (symmetric, applied to BOTH revenue and payout)"
    return f"(1+(({inp('car_value')}/{REF_CAR})-1)*{inp('prestige_coeff')})"

def xabofac():
    "ABO uplift factor = 1 + narzut%/100 (kryje oklejenie + logistykę; ABO droższy od SUB)"
    return f"(1+{inp('abo_uplift')}/100)"

def xroute_driver():
    "Blended driver rate = miejska×(1−hw) + pozamiejska×hw  (hw = udział km pozamiejskich)"
    hw = f"({inp('highway_share')}/100)"
    return f"({inp('driver_rate_per_km')}*(1-{hw})+{inp('driver_rate_highway')}*{hw})"

def xroute_adv():
    "Blended advertiser rate = miejska×(1−hw) + pozamiejska×hw"
    hw = f"({inp('highway_share')}/100)"
    return f"({inp('adv_rate_per_km')}*(1-{hw})+{inp('adv_rate_highway')}*{hw})"

def xrev(is_abo=False):
    "Revenue per car per month = km × blended_adv_rate × prestige × (narzut ABO jeśli abo)"
    base = f"{inp('km_per_month')}*{xroute_adv()}*{xprestige()}"
    return f"{base}*{xabofac()}" if is_abo else base

def xdriver():
    "Driver payout per car per month = km × blended_driver_rate × prestige × wrap_factor (identyczny SUB/ABO)"
    return f"{inp('km_per_month')}*{xroute_driver()}*{xprestige()}*{inp('wrap_factor')}"

def xabo_months():
    "ABO: minimalny abonament 12 miesięcy → podłoga amortyzacji oklejenia"
    return f"MAX({inp('months')},12)"

def xwrap():
    "Wrap amort per car per month (abo only) — podłoga 12 mies."
    return f"{inp('wrap_cost')}/{xabo_months()}"

def xunwrap():
    "Unwrap amort per car per month (abo only) — podłoga 12 mies."
    return f"{inp('unwrap_cost')}/{xabo_months()}"

def xvarCost(is_abo):
    d = xdriver()
    if is_abo:
        return f"{d}+{xwrap()}+{xunwrap()}"
    return d

def xgrossMargin(is_abo):
    return f"({xrev(is_abo)})-({xvarCost(is_abo)})"

def xtotalRev(n, is_abo=False):
    "n = reference string or integer"
    return f"({xrev(is_abo)})*{n}"

def xtotalVar(n, is_abo):
    return f"({xvarCost(is_abo)})*{n}"

def xnetProfit(n, is_abo):
    "Monthly net profit = totalRev - totalVar - opsCost  (legal is one-time, not monthly)"
    return f"=({xtotalRev(n, is_abo)})-({xtotalVar(n, is_abo)})-{inp('ops_cost')}"

def xbepMonths(n, is_abo):
    "Months to recover one-time legal cost"
    net = f"({xtotalRev(n, is_abo)})-({xtotalVar(n, is_abo)})-{inp('ops_cost')}"
    return f"=IFERROR(IF(({net})>0,{inp('legal_cost')}/({net}),\"N/A\"),\"N/A\")"

def xadvTotal(n, months_ref, is_abo):
    monthly = f"({xrev(is_abo)})*{n}*{months_ref}"
    if is_abo:
        return f"={monthly}"
    wrap_total = f"({inp('wrap_cost')}+{inp('unwrap_cost')})*{n}"
    return f"=({monthly})+({wrap_total})"

# ──────────────────────────────────────────────────────────────
# CREATE MODEL SHEET (Subskrypcyjny / Abonamentowy)
# ──────────────────────────────────────────────────────────────
def make_model_sheet(wb, title, is_abo, thm, thm_bg):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCDEF', [34, 22, 18, 18, 18, 18]):
        ws.column_dimensions[col].width = w

    NC = inp('num_cars'); MT = inp('months'); OPS = inp('ops_cost'); LGL = inp('legal_cost')

    # Title
    ws.merge_cells('A1:F1')
    c = ws['A1']; c.value = f"AdRide — {title}"
    c.fill = fill('0F3460'); c.font = Font(bold=True, color='FFFFFF', size=13, name='Calibri')
    c.alignment = aln('center'); ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2'); c = ws['A2']
    desc = ("Abonamentowy — AdRide organizuje oklejenie. Koszt absorbowany w marży."
            if is_abo else
            "Subskrypcyjny — Reklamodawca organizuje oklejenie. AdRide pobiera fee platformy.")
    c.value = f"  {'✦' if is_abo else '◆'}  {desc}"
    c.fill = fill(thm_bg); c.font = Font(italic=True, size=9, name='Calibri')
    c.alignment = aln('left', 'center'); ws.row_dimensions[2].height = 18

    row = 4

    # ── Składowe formuły ──
    sec(ws, row, "📐 Formuły składowe (/ auto / mies.)", 6, bg=thm); row += 1
    hdr(ws, row, ["Parametr", "Formuła / opis", "Wartość", "", "", ""], bg='1A2744', size=9); row += 1

    rows_def = [
        ("Współczynnik wartości auta",      f"=({inp('car_value')}/{REF_CAR})", NUM2,  None),
        ("Współczynnik prestiżu (symetr.)", f"={xprestige()}", '0.000', None),
        ("Ekspozycja: km / auto / mies.",   f"={inp('km_per_month')}", NUM, None),
        ("Stawka kier. mieszana / km",      f"={xroute_driver()}", NUM2, None),
        ("Stawka rekl. mieszana / km",      f"={xroute_adv()}", NUM2, None),
        ("Wypłata kierowcy / auto / mies.", f"={xdriver()}", PLN, 'F8D7DA'),
        ("Przychód AdRide / auto / mies.",  f"={xrev(is_abo)}", PLN, 'D4EDDA'),
    ]
    if is_abo:
        rows_def += [
            ("Amortyzacja oklejenia / auto / mies.", f"={xwrap()}", PLN, 'F8D7DA'),
            ("Amortyzacja demontażu / auto / mies.", f"={xunwrap()}", PLN, 'F8D7DA'),
        ]
    rows_def += [
        ("Koszt zmienny / auto / mies.",    f"={xvarCost(is_abo)}", PLN, 'F8D7DA'),
        ("Marża brutto / auto / mies.",     f"={xgrossMargin(is_abo)}", PLN, thm_bg),
        ("Marża brutto %",
         f"=IFERROR(({xgrossMargin(is_abo)})/({xrev(is_abo)})*100,0)", PCT, thm_bg),
    ]

    for label, formula, fmt, bg_ in rows_def:
        ws.row_dimensions[row].height = 19
        sc(ws, row, 1, label, bg='EEF2FF')
        sc(ws, row, 2, "", bg='F8F8F8', italic=True, fg='666666', size=9)
        sc(ws, row, 3, formula, bg=bg_ or 'FFFFFF', fmt=fmt, h='right', bold=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        row += 1

    row += 1

    # ── Waterfall P&L (miesięczny, bez kosztu prawnego!) ──
    sec(ws, row, "💧 Waterfall P&L platformy / mies.  (audyt prawny = koszt jednorazowy, poza P&L)", 6, bg=thm)
    row += 1
    hdr(ws, row, ["Pozycja", "Opis", "Kwota (zł/mies.)", "% przychodów", "", ""], bg='1A2744', size=9); row += 1

    wf_items = [
        ("📈 Przychód (fee platformy)", f"={NC} aut × rev/auto", f"=({xtotalRev(NC, is_abo)})", 'D4EDDA', '155724'),
        ("📉 Wypłata kierowców",        f"={NC} aut × payout/auto", f"=-({xdriver()})*{NC}", 'F8D7DA', '721C24'),
    ]
    if is_abo:
        wf_items += [
            ("📉 Oklejenie (amort.)",   f"={NC} aut × wrapCost/months", f"=-({xwrap()})*{NC}", 'F8D7DA', '721C24'),
            ("📉 Demontaż (amort.)",    f"={NC} aut × unwrapCost/months", f"=-({xunwrap()})*{NC}", 'F8D7DA', '721C24'),
        ]
    wf_items += [
        ("📉 Koszty ops. platformy",    "stały / mies.", f"=-{OPS}", 'F8D7DA', '721C24'),
    ]

    total_rev_val = f"({xtotalRev(NC, is_abo)})"
    wf_rows = []
    for label, desc, formula, bg_, fg_ in wf_items:
        ws.row_dimensions[row].height = 19
        sc(ws, row, 1, label, bg=bg_)
        sc(ws, row, 2, desc, bg='F8F8F8', italic=True, fg='666666', size=9)
        sc(ws, row, 3, formula, bg=bg_, fg=fg_, fmt=PLN, h='right', bold=True)
        sc(ws, row, 4, f"=IFERROR({formula.lstrip('=')}/({total_rev_val})*100,0)", bg=bg_, fg=fg_, fmt=PCT, h='right')
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        row += 1

    # Net profit row
    net_formula = xnetProfit(NC, is_abo)
    ws.row_dimensions[row].height = 24
    sc(ws, row, 1, "✅ ZYSK OPERACYJNY / mies.", bg='16213E', fg='FFFFFF', bold=True)
    sc(ws, row, 2, "bez kosztu prawnego (jednorazowy)", bg='16213E', fg='8B949E', italic=True, size=9)
    sc(ws, row, 3, net_formula, bg='16213E', fg='00D4AA', fmt=PLN, h='right', bold=True, size=12)
    sc(ws, row, 4, f"=IFERROR(({net_formula[1:]})/({total_rev_val})*100,0)", bg='16213E', fg='00D4AA', fmt=PCT, h='right', bold=True)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    row += 2

    # ── Jednorazowy koszt startu i BEP ──
    sec(ws, row, "🎯 Jednorazowy koszt startu i Break-Even", 6, bg=thm); row += 1
    hdr(ws, row, ["Metryka", "Wartość", "Interpretacja", "", "", ""], bg='1A2744', size=9); row += 1

    bep_items = [
        ("Jednorazowy audyt prawny", f"={LGL}", PLN, "Koszt poniesiony RAZ przy starcie — niezależny od liczby umów"),
        ("Zysk operacyjny / mies.",  f"=({net_formula[1:]})", PLN, "Miesięczna nadwyżka do spłaty startu"),
        ("Miesięcy do zwrotu audytu", xbepMonths(NC, is_abo)[1:], '0.0',
         f"= audyt prawny ÷ zysk operacyjny / mies. (przy {NC} autach)"),
        ("Pokrycie kosztów stałych", f"=IFERROR({OPS}/({xtotalRev(NC, is_abo)})-({xtotalVar(NC,is_abo)}),\"N/A\")", '0.0%',
         "% zysku brutto zajęty przez ops (< 1 = opłacalne operacyjnie)"),
    ]

    for label, formula, fmt_, note in bep_items:
        ws.row_dimensions[row].height = 20
        sc(ws, row, 1, label, bg='EEF2FF')
        formula_val = formula if formula.startswith('=') else f"={formula}"
        sc(ws, row, 2, formula_val, bg=thm_bg, fmt=fmt_, h='right', bold=True)
        sc(ws, row, 3, note, bg='F8F8F8', italic=True, fg='666666', size=9, wrap=True)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1

    # ── Koszt reklamodawcy ──
    sec(ws, row, "💼 Całkowity koszt reklamodawcy (za kampanię)", 6, bg=thm); row += 1
    hdr(ws, row, ["Składnik", "Wzór", "Łącznie", "/ auto", "/ auto / mies.", ""], bg='1A2744', size=9); row += 1

    monthly_total = f"=({xrev(is_abo)})*{NC}*{MT}"
    sc(ws, row, 1, "Opłata platformy AdRide", bg='EEF2FF')
    sc(ws, row, 2, "rev/auto × N × months", bg='F8F8F8', italic=True, fg='666666', size=9)
    sc(ws, row, 3, monthly_total, bg='FFFFFF', fmt=PLN, h='right', bold=True)
    sc(ws, row, 4, f"=({xrev(is_abo)})*{MT}", bg='FFFFFF', fmt=PLN, h='right')
    sc(ws, row, 5, f"={xrev(is_abo)}", bg='FFFFFF', fmt=PLN, h='right')
    sc(ws, row, 6, "", bg='FFFFFF'); row += 1

    if is_abo:
        sc(ws, row, 1, "Koszt oklejenia + demontażu", bg='EEF2FF')
        sc(ws, row, 2, "Absorbowany w marży AdRide", bg='D4EDDA', italic=True, fg='155724', size=9)
        for c_ in range(3, 7):
            sc(ws, row, c_, 0 if c_ < 6 else "", bg='D4EDDA', fg='155724', fmt=PLN if c_ < 6 else None, h='right')
        ws.cell(row=row, column=6).value = "✅ Wliczone w usługę"
        ws.cell(row=row, column=6).font = Font(bold=True, color='155724', size=9, name='Calibri')
        adv_total = monthly_total
    else:
        wrap_t = f"=({inp('wrap_cost')}+{inp('unwrap_cost')})*{NC}"
        sc(ws, row, 1, "Koszt oklejenia + demontażu (reklamodawca płaci sam)", bg='FFF3CD')
        sc(ws, row, 2, "(wrapCost+unwrapCost)×N — poza AdRide", bg='FFF3CD', italic=True, fg='7D6608', size=9)
        sc(ws, row, 3, wrap_t, bg='FFF3CD', fg='7D6608', fmt=PLN, h='right', bold=True)
        sc(ws, row, 4, f"={inp('wrap_cost')}+{inp('unwrap_cost')}", bg='FFF3CD', fg='7D6608', fmt=PLN, h='right')
        sc(ws, row, 5, f"=({inp('wrap_cost')}+{inp('unwrap_cost')})/{MT}", bg='FFF3CD', fg='7D6608', fmt=PLN, h='right')
        ws.cell(row=row, column=6).value = "⚠ Poza platformą"
        adv_total = f"=({monthly_total[1:]})+({wrap_t[1:]})"

    row += 1
    ws.row_dimensions[row].height = 22
    sc(ws, row, 1, "💰 ŁĄCZNY KOSZT REKLAMODAWCY", bg='16213E', fg='FFFFFF', bold=True)
    sc(ws, row, 2, "", bg='16213E')
    sc(ws, row, 3, adv_total if adv_total.startswith('=') else f"={adv_total}", bg='16213E', fg='58A6FF', fmt=PLN, h='right', bold=True, size=12)
    adv_expr = adv_total.lstrip('=')
    sc(ws, row, 4, f"=({adv_expr})/{NC}", bg='16213E', fg='58A6FF', fmt=PLN, h='right', bold=True)
    sc(ws, row, 5, f"=({adv_expr})/({NC}*{MT})", bg='16213E', fg='58A6FF', fmt=PLN, h='right', bold=True)
    sc(ws, row, 6, "", bg='16213E')
    row += 2

    # ── P&L / auto ──
    sec(ws, row, "🔍 P&L na jedno auto / mies.", 6, bg=thm); row += 1
    hdr(ws, row, ["Pozycja", "Kwota / auto", "", "", "", ""], bg='1A2744', size=9); row += 1

    unit_items = [
        ("Przychód (fee platformy)",    f"={xrev(is_abo)}",                    PLN, 'D4EDDA', '155724', False),
        ("Wypłata kierowcy",            f"=-({xdriver()})",                    PLN, 'F8D7DA', '721C24', False),
    ]
    if is_abo:
        unit_items += [
            ("Oklejenie (amort.)",      f"=-({xwrap()})",                      PLN, 'F8D7DA', '721C24', False),
            ("Demontaż (amort.)",       f"=-({xunwrap()})",                    PLN, 'F8D7DA', '721C24', False),
        ]
    unit_items += [
        ("Marża brutto",                f"={xgrossMargin(is_abo)}",            PLN, thm_bg,   '000000', True),
        ("Ops. platforma / auto",       f"=-{OPS}/{NC}",                       PLN, 'F8D7DA', '721C24', False),
        ("Marża operacyjna netto / auto",f"=({xnetProfit(NC,is_abo)[1:]})/{NC}", PLN, '16213E', '00D4AA', True),
    ]

    for label, formula, fmt_, bg_, fg_, bold_ in unit_items:
        ws.row_dimensions[row].height = 20 if bold_ else 18
        sc(ws, row, 1, label, bg='EEF2FF' if bg_ == '16213E' else bg_)
        sc(ws, row, 2, formula, bg=bg_, fg=fg_, fmt=fmt_, h='right', bold=bold_,
           size=11 if bg_ == '16213E' else 10)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        row += 1

    return ws


ws_sub = make_model_sheet(wb, "Subskrypcyjny", False, '1A7A5E', 'E8FDF5')
ws_abo = make_model_sheet(wb, "Abonamentowy",  True,  '1A4A7A', 'E8F4FD')

# ──────────────────────────────────────────────────────────────
# SHEET 4 — PORÓWNANIE
# ──────────────────────────────────────────────────────────────
ws_cmp = wb.create_sheet("Porównanie modeli")
ws_cmp.sheet_view.showGridLines = False
for col, w in zip('ABCD', [36, 22, 22, 30]): ws_cmp.column_dimensions[col].width = w

ws_cmp.merge_cells('A1:D1')
c = ws_cmp['A1']; c.value = "AdRide — Porównanie: Subskrypcyjny vs Abonamentowy"
c.fill = fill('0F3460'); c.font = Font(bold=True, color='FFFFFF', size=13, name='Calibri')
c.alignment = aln('center'); ws_cmp.row_dimensions[1].height = 30

NC = inp('num_cars'); MT = inp('months')

hdr(ws_cmp, 2,
    ["Metryka",
     "Subskrypcyjny\n(reklamodawca org. oklejenie)",
     "Abonamentowy\n(AdRide org. oklejenie)",
     "Uwagi"],
    bg='0F3460', size=10)
ws_cmp.row_dimensions[2].height = 32

cmp_data = [
    (None, "PRZYCHODY ADRIDE"),
    ("Przychód / auto / mies.", f"={xrev(False)}", f"={xrev(True)}", "ABO: +narzut"),
    ("Przychód łącznie / mies.", f"={xtotalRev(NC,False)}", f"={xtotalRev(NC,True)}", "ABO wyższy o narzut"),

    (None, "KOSZTY ADRIDE"),
    ("Payout kierowcy / auto", f"={xdriver()}", f"={xdriver()}", "Identyczny"),
    ("Koszt oklejenia / auto / mies.", "0 zł", f"={xwrap()}", "Abo: amortyzowane"),
    ("Koszt demontażu / auto / mies.", "0 zł", f"={xunwrap()}", "Abo: amortyzowane"),
    ("Koszt zmienny / auto",    f"={xvarCost(False)}", f"={xvarCost(True)}", "Abo wyższy"),

    (None, "MARŻE ADRIDE (miesięczne)"),
    ("Marża brutto / auto",     f"={xgrossMargin(False)}", f"={xgrossMargin(True)}", "Sub wyższa o koszt oklejenia"),
    ("Marża brutto %",
     f"=IFERROR(({xgrossMargin(False)})/({xrev(False)})*100,0)",
     f"=IFERROR(({xgrossMargin(True)})/({xrev(True)})*100,0)", "%"),
    ("Zysk operacyjny / mies.", xnetProfit(NC,False), xnetProfit(NC,True), "Bez kosztu prawnego"),

    (None, "KOSZT JEDNORAZOWY (niezależny od modelu)"),
    ("Audyt prawny (jednorazowo)", f"={inp('legal_cost')}", f"={inp('legal_cost')}", "Płacony raz — niezależnie od liczby klientów"),
    ("Mies. do zwrotu audytu", xbepMonths(NC,False), xbepMonths(NC,True), "Przy aktualnych ustawieniach"),

    (None, "KOSZTY REKLAMODAWCY"),
    ("Opłata AdRide (łącznie)", f"=({xrev(False)})*{NC}*{MT}", f"=({xrev(True)})*{NC}*{MT}", "ABO wyższy o narzut"),
    ("Koszt oklejenia (kto?)",  f"=({inp('wrap_cost')}+{inp('unwrap_cost')})*{NC}", "0 zł (w usłudze)", "Sub: reklamodawca płaci"),
    ("ŁĄCZNY KOSZT REKLAMODAWCY",
     f"=({xrev(False)})*{NC}*{MT}+({inp('wrap_cost')}+{inp('unwrap_cost')})*{NC}",
     f"=({xrev(True)})*{NC}*{MT}",
     "Porównaj który tańszy dla reklamodawcy"),
]

row_cmp = 3
for item in cmp_data:
    if item[0] is None:
        sec(ws_cmp, row_cmp, item[1], 4, bg='E94560'); row_cmp += 1; continue

    label, val_s, val_a, note = item
    ws_cmp.row_dimensions[row_cmp].height = 20
    sc(ws_cmp, row_cmp, 1, label, bg='EEF2FF')

    for col_i, val, bg_ in [(2, val_s, 'E8FDF5'), (3, val_a, 'E8F4FD')]:
        c = ws_cmp.cell(row=row_cmp, column=col_i)
        if isinstance(val, str) and (val.startswith('=') or val == "0 zł" or val == "0 zł (w usłudze)"):
            c.value = val if val.startswith('=') else val
            if val.startswith('='):
                if "100" in label or "%" in label: c.number_format = PCT
                else: c.number_format = PLN
        else:
            c.value = val
        c.fill = fill(bg_); c.font = fnt(bold=True, size=10)
        c.alignment = aln('right', 'center'); c.border = brd()

    sc(ws_cmp, row_cmp, 4, note, bg='F8F8F8', italic=True, fg='666666', size=9, wrap=True)
    row_cmp += 1

# ──────────────────────────────────────────────────────────────
# SHEET 5 — SCENARIUSZE
# ──────────────────────────────────────────────────────────────
ws_sc = wb.create_sheet("Scenariusze")
ws_sc.sheet_view.showGridLines = False
for i, w in enumerate([5,18,18,14,14,4,18,14,14,18,14], 1):
    ws_sc.column_dimensions[get_column_letter(i)].width = w

ws_sc.merge_cells('A1:K1')
c = ws_sc['A1']; c.value = "AdRide — Scenariusze: Wpływ liczby aut (zysk operacyjny, bez kosztu prawnego)"
c.fill = fill('0F3460'); c.font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
c.alignment = aln('center'); ws_sc.row_dimensions[1].height = 26

hdr(ws_sc, 2, ["Aut","Rev SUB/mc","Zysk SUB","M%SUB","BEP SUB","","Zysk ABO","M%ABO","BEP ABO","Δ SUB−ABO","Lepszy?"],
    bg='0F3460', size=9)

for s_row, n in enumerate([1,3,5,10,15,20,30,50,75,100], 3):
    ws_sc.row_dimensions[s_row].height = 18
    alt = 'F4F6FF' if s_row % 2 == 0 else 'FFFFFF'
    net_sub = f"({xtotalRev(n,False)})-({xtotalVar(n,False)})-{inp('ops_cost')}"
    net_abo = f"({xtotalRev(n,True)})-({xtotalVar(n,True)})-{inp('ops_cost')}"
    bep_s = f'=IFERROR(IF(({net_sub})>0,{inp("legal_cost")}/({net_sub}),"N/A"),"N/A")'
    bep_a = f'=IFERROR(IF(({net_abo})>0,{inp("legal_cost")}/({net_abo}),"N/A"),"N/A")'
    vals = [n, f"={xtotalRev(n,False)}", f"={net_sub}", f"=IFERROR(({net_sub})/({xtotalRev(n,False)})*100,0)", bep_s,
            "", f"={net_abo}", f"=IFERROR(({net_abo})/({xtotalRev(n,True)})*100,0)", bep_a,
            f"=({net_sub})-({net_abo})", f'=IF(({net_sub})>({net_abo}),"SUB","ABO")']
    fmts = [NUM, PLN, PLN, PCT, '0.0', None, PLN, PCT, '0.0', PLN, '@']
    bgs  = [alt,alt,'E8FDF5','E8FDF5','E8FDF5',alt,'E8F4FD','E8F4FD','E8F4FD',alt,alt]
    for ci, (v, f_, b_) in enumerate(zip(vals, fmts, bgs), 1):
        c = ws_sc.cell(row=s_row, column=ci)
        c.value = v; c.fill = fill(b_)
        c.font = fnt(bold=(ci==1), size=10)
        c.alignment = aln('right' if ci>1 else 'center', 'center')
        if f_: c.number_format = f_
        c.border = brd()

# ──────────────────────────────────────────────────────────────
# SHEET 6 — WRAŻLIWOŚĆ
# ──────────────────────────────────────────────────────────────
ws_sv = wb.create_sheet("Wrażliwość — km na mies.")
ws_sv.sheet_view.showGridLines = False
for i, w in enumerate([18,18,18,18,14,4,18,18,14], 1):
    ws_sv.column_dimensions[get_column_letter(i)].width = w

ws_sv.merge_cells('A1:I1')
c = ws_sv['A1']; c.value = "AdRide — Wrażliwość na km / mies. (marża SUB vs ABO, przy aktualnej wartości auta)"
c.fill = fill('0F3460'); c.font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
c.alignment = aln('center'); ws_sv.row_dimensions[1].height = 26

ws_sv.merge_cells('A2:E2')
ws_sv['A2'].value = "SUBSKRYPCYJNY (brak kosztów oklejenia dla AdRide)"
ws_sv['A2'].fill = fill('1A7A5E'); ws_sv['A2'].font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
ws_sv['A2'].alignment = aln('center')
ws_sv.merge_cells('G2:I2')
ws_sv['G2'].value = "ABONAMENTOWY (wrapAmort wliczone w marżę)"
ws_sv['G2'].fill = fill('1A4A7A'); ws_sv['G2'].font = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
ws_sv['G2'].alignment = aln('center')
ws_sv.row_dimensions[2].height = 20

hdr(ws_sv, 3,
    ["km / mies.","Payout kierowcy","Przychód platf.","Marża brutto","Marża %",
     "","Payout kierowcy","Marża brutto (z okl.)","Marża %"],
    bg='1A2744', size=9)

for sv_row, km in enumerate([200,500,1000,2000,3000,5000,7500,10000], 4):
    ws_sv.row_dimensions[sv_row].height = 18
    alt = 'F4F6FF' if sv_row % 2 == 0 else 'FFFFFF'
    # Expressions for fixed km value (car value & prestiż z aktualnych suwaków)
    dp  = f"{km}*{xroute_driver()}*{xprestige()}*{inp('wrap_factor')}"
    rev_sub = f"{km}*{xroute_adv()}*{xprestige()}"
    rev_abo = f"({rev_sub})*{xabofac()}"
    gm_sub = f"({rev_sub})-({dp})"
    gm_abo = f"({rev_abo})-({dp})-{inp('wrap_cost')}/{xabo_months()}-{inp('unwrap_cost')}/{xabo_months()}"
    row_vals = [
        km, f"={dp}", f"={rev_sub}", f"={gm_sub}", f"=IFERROR(({gm_sub})/({rev_sub})*100,0)",
        "", f"={dp}", f"={gm_abo}", f"=IFERROR(({gm_abo})/({rev_abo})*100,0)",
    ]
    row_fmts = [NUM, PLN, PLN, PLN, PCT, None, PLN, PLN, PCT]
    row_bgs  = [alt,'F8D7DA','D4EDDA','E8FDF5','E8FDF5', alt,'F8D7DA','E8F4FD','E8F4FD']
    for ci, (v, f_, b_) in enumerate(zip(row_vals, row_fmts, row_bgs), 1):
        c = ws_sv.cell(row=sv_row, column=ci)
        c.value = v; c.fill = fill(b_); c.border = brd()
        c.font = fnt(bold=(ci==1), size=10)
        c.alignment = aln('right', 'center')
        if f_: c.number_format = f_

note_row = 13
ws_sv.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
c = ws_sv.cell(row=note_row, column=1)
c.value = ("⚠ MODEL v2 (per-km): przychód = km × stawka_rekl. × prestiż; payout = km × stawka_kier. × prestiż × oklejenie. "
           "Prestiż symetryczny (ten sam wsp. 0.3 po obu stronach → marża % stała). "
           "Abo: wrapAmort+unwrapAmort w marży; audyt prawny jednorazowy poza P&L.")
c.fill = fill('FFF3CD'); c.font = Font(italic=True, color='7D6608', size=9, name='Calibri')
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws_sv.row_dimensions[note_row].height = 36

# ──────────────────────────────────────────────────────────────
# SAVE
# ──────────────────────────────────────────────────────────────
sheet_order = ["Dane wejściowe","Subskrypcyjny","Abonamentowy","Porównanie modeli","Scenariusze","Wrażliwość — km na mies."]
for i, name in enumerate(sheet_order):
    wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

OUTPUT = r"C:\Users\andrz\Downloads\AdRide\adride_unit_economics.xlsx"
wb.save(OUTPUT)
print(f"XLSX saved: {OUTPUT}")

# ──────────────────────────────────────────────────────────────
# ADD SCROLL BARS via VML injection into the XLSX ZIP
# ──────────────────────────────────────────────────────────────
def inject_vml_scrollbars(xlsx_path, scroll_cfg):
    """
    Insert horizontal form-control scroll bars into sheet1 (Dane wejściowe).
    scroll_cfg: key -> (excel_row_1indexed, min_sb, max_sb, scale, val_sb)
      scale==1 : scroll bar linked directly to $B$row
      scale!=1 : scroll bar linked to $F$row (helper); B already has =F*scale formula
                 (scale może być ułamkowa, np. 0.05 dla stawek zł/km)
    """
    # Approximate pixel dimensions at 96 DPI, Calibri 11pt
    ROW_H_PX   = 29   # 22pt row height
    SB_H_PX    = 16   # scroll bar height
    E_W_PX     = 100  # column E ≈ 14 chars ≈ 100px
    TOP_OFF    = (ROW_H_PX - SB_H_PX) // 2
    BOT_OFF    = TOP_OFF + SB_H_PX

    shapes = []
    for sid, (key, (sb_row, min_sb, max_sb, scale, val_sb)) in enumerate(
            sorted(scroll_cfg.items(), key=lambda x: x[1][0]), start=1025):
        row_idx = sb_row - 1  # 0-indexed for VML anchor
        linked  = f"$B${sb_row}" if scale == 1 else f"$F${sb_row}"
        anchor  = f"4,2,{row_idx},{TOP_OFF},4,{E_W_PX - 2},{row_idx},{BOT_OFF}"
        page    = max(1, (max_sb - min_sb) // 10)
        shapes.append(f"""<v:shape id="_x0000_s{sid}" type="#_x0000_t201"
   style='position:absolute;margin-left:0;margin-top:0;width:72pt;height:12pt;z-index:{sid-1024};mso-position-horizontal-relative:cells;mso-position-vertical-relative:cells'
   o:insetmode="auto">
  <v:fill color2="#c0c0c0" o:detectmouseclick="t"/>
  <v:stroke dashstyle="solid" o:ext="view"/>
  <o:lock v:ext="edit" rotation="t"/>
  <x:ClientData ObjectType="Scroll">
   <x:Anchor>{anchor}</x:Anchor>
   <x:PrintObject>False</x:PrintObject>
   <x:AutoFill>False</x:AutoFill>
   <x:DX>15</x:DX>
   <x:Inc>1</x:Inc>
   <x:Page>{page}</x:Page>
   <x:Min>{int(min_sb)}</x:Min>
   <x:Max>{int(max_sb)}</x:Max>
   <x:Val>{int(val_sb)}</x:Val>
   <x:Horiz/>
   <x:FmlaLink>{linked}</x:FmlaLink>
  </x:ClientData>
 </v:shape>""")

    vml = (
        '<xml xmlns:v="urn:schemas-microsoft-com:vml"'
        ' xmlns:o="urn:schemas-microsoft-com:office:office"'
        ' xmlns:x="urn:schemas-microsoft-com:office:excel">\n'
        ' <o:shapelayout v:ext="edit"><o:idmap v:ext="edit" data="1"/></o:shapelayout>\n'
        ' <v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201"'
        ' path="m,l,21600,21600,21600,21600,xe">'
        '<v:stroke joinstyle="miter"/>'
        '<v:path shadowok="f" o:extrusionok="f" gradientshapeok="t" o:connecttype="rect"/>'
        '</v:shapetype>\n'
        + '\n'.join(shapes)
        + '\n</xml>'
    )

    VML_PATH  = 'xl/drawings/vmlDrawing1.vml'
    RELS_PATH = 'xl/worksheets/_rels/sheet1.xml.rels'
    REL_ENTRY = (b'<Relationship Id="rId_vml1"'
                 b' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing"'
                 b' Target="../drawings/vmlDrawing1.vml"/>')
    CT_ENTRY  = (b'<Override PartName="/xl/drawings/vmlDrawing1.vml"'
                 b' ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>')

    tmp = xlsx_path + '.tmp'
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        existing = {i.filename for i in zin.infolist()}
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == '[Content_Types].xml':
                    if b'vmlDrawing' not in data:
                        data = data.replace(b'</Types>', CT_ENTRY + b'</Types>')

                elif item.filename == 'xl/worksheets/sheet1.xml':
                    if b'legacyDrawing' not in data:
                        # Ensure r: namespace is declared (openpyxl omits it when no drawings exist)
                        if b'xmlns:r=' not in data[:600]:
                            data = data.replace(
                                b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
                                b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
                                b' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">',
                            )
                        data = data.replace(b'</worksheet>',
                            b'<legacyDrawing r:id="rId_vml1"/></worksheet>')

                elif item.filename == RELS_PATH:
                    if b'vmlDrawing' not in data:
                        data = data.replace(b'</Relationships>', REL_ENTRY + b'</Relationships>')

                zout.writestr(item, data)

            zout.writestr(VML_PATH, vml.encode('utf-8'))

            if RELS_PATH not in existing:
                rels_xml = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n'
                    ' <Relationship Id="rId_vml1"'
                    ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing"'
                    ' Target="../drawings/vmlDrawing1.vml"/>\n'
                    '</Relationships>'
                )
                zout.writestr(RELS_PATH, rels_xml.encode('utf-8'))

    shutil.move(tmp, xlsx_path)
    print(f"VML scroll bars injected: {len(shapes)} controls")


inject_vml_scrollbars(OUTPUT, scroll_config)
print("Gotowe!")
